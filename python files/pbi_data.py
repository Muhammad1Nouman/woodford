import mysql.connector
from mysql.connector import Error
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.utils import column_index_from_string

# Load workbook once
wb = load_workbook(r'D:\Power BI\woodfordbros.com\Woodford All - Power BI Files\PBI Data.xlsx', data_only=True)

def get_database_connection():
    try:
        # Connect to the MySQL database
        connection = mysql.connector.connect(
            host='10.0.0.75',
            port=3306,
            user='pbiRefresh',
            password='pbiRefresh',
            database='pre_stage'
        )
        cursor = connection.cursor()
        print('Database connection successful')
        return cursor, connection
    except mysql.connector.Error as error:
        print("Error connecting to MySQL database:", error)
        return None, None

# Function to extract data from Excel and insert into MySQL
def extract_and_insert(sheet_name, db_column):
    ws = wb[sheet_name]
    cursor, connection = get_database_connection()

    if cursor is None or connection is None:
        return

    print(f"Processing {sheet_name} sheet...")

    row = 9
    while True:
        year = ws.cell(row=row, column=1).value
        if not isinstance(year, int):
            break

        for col in range(9, 21):  # Columns I to T
            month = col - 8
            raw_value = ws.cell(row=row, column=col).value

            if isinstance(raw_value, (int, float)):
                formatted_value = format(raw_value, '.10f')
            else:
                formatted_value = raw_value

            if formatted_value:
                print(f"Year {year}, Month {month}: {formatted_value}")
                cursor.execute(f"""
                    INSERT INTO monthly_distribution (year, month, {db_column})
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE {db_column} = VALUES({db_column})
                """, (year, month, formatted_value))

        row += 1

    connection.commit()
    print(f"✅ Insertion successful for {sheet_name}.")
    cursor.close()
    connection.close()
    print("🔒 MySQL connection closed.")

# Extracting sales teams
def extract_sales_teams(sheet_name, table_name, col_1, col_2):
    ws = wb[sheet_name]
    cursor, connection = get_database_connection()

    if cursor is None or connection is None:
        return

    print(f"Processing {sheet_name},{table_name} sheet...")
    row = 2  # Assuming row 1 has headers
    while True:
        team_name = ws.cell(row=row, column=1).value
        employee_name = ws.cell(row=row, column=2).value

        # Stop if both cells are empty (end of data)
        if not team_name and not employee_name:
            break

        if team_name and employee_name:
            print(f"Inserting: Team = {team_name}, Employee = {employee_name}")
            cursor.execute(f"""
                INSERT IGNORE INTO pre_stage.{table_name} ({col_1}, {col_2})
                VALUES (%s, %s)
            """, (team_name, employee_name))

        row += 1

    connection.commit()
    print("✅ Sales teams inserted successfully.")
    cursor.close()
    connection.close()
    print("🔒 MySQL connection closed.")

def extract_and_insert_goals_v3(sheet_name, col_name, col_start, col_end):
    ws = wb[sheet_name]
    cursor, connection = get_database_connection()

    if cursor is None or connection is None:
        return

    print(f"Processing {sheet_name} sheet for column: {col_name}...")

    start_col = column_index_from_string(col_start)
    end_col = column_index_from_string(col_end)

    # Check if row 3 has any non-empty cells in the range
    row3_has_data = any(ws.cell(row=3, column=col).value not in (None, "") for col in range(start_col, end_col + 1))

    # Build list of tuples based on row 2 and conditionally row 3
    categories = []
    for col in range(start_col, end_col + 1):
        category = ws.cell(row=2, column=col).value
        name = ws.cell(row=3, column=col).value if row3_has_data else None
        categories.append((col, category, name))

    row = 4
    while True:
        year = ws.cell(row=row, column=1).value
        if year is None:
            break

        for col, category, name in categories:
            goal = ws.cell(row=row, column=col).value
            if year and category and goal not in (None, ""):
                if name:
                    print(f"Year {year}, Category {category}, Name {name}: {col_name} = {goal}")
                else:
                    print(f"Year {year}, Category {category}: {col_name} = {goal}")

                # You can choose to include name in DB only if it's not None
                # Example:
                if name:
                    cursor.execute(f"""
                        INSERT INTO goal_distribution_name(year, product_category, name, {col_name})
                        VALUES (%s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE {col_name} = VALUES({col_name})
                    """, (year, category, name, goal))
                else:
                    cursor.execute(f"""
                        INSERT INTO goal_distribution(year, product_category, {col_name})
                        VALUES (%s, %s, %s)
                        ON DUPLICATE KEY UPDATE {col_name} = VALUES({col_name})
                    """, (year, category, goal))

        row += 1

    connection.commit()
    print("Data insertion completed.")

def extract_wsg_goal_tracker():
    ws = wb["WSG Goal Tracker"]
    cursor, connection = get_database_connection()

    if cursor is None or connection is None:
        return

    print("Processing WSG Goal Tracker sheet...")

    # TRUNCATE the table before inserting new data
    try:
        cursor.execute("TRUNCATE TABLE pre_stage.wsg_goal_tracker")
        print("✅ Table pre_stage.wsg_goal_tracker truncated successfully.")
    except Error as e:
        print(f"Error truncating table: {e}")
        connection.rollback()
        cursor.close()
        connection.close()
        return

    # Extract values from cells K12, S14, and AI12
    k_col = column_index_from_string("K")
    s_col = column_index_from_string("S")
    ai_col = column_index_from_string("AI")
    
    sales_wsg_percent = ws.cell(row=12, column=k_col).value
    net_sales = ws.cell(row=14, column=s_col).value
    sales_production_wsg = ws.cell(row=12, column=ai_col).value  # New value from AI12
    
    # Print the extracted values for verification
    print(f"Raw values from Excel - Sales WSG%: {sales_wsg_percent}, Net Sales: {net_sales}, Sales-Production WSG: {sales_production_wsg}")
    
    # Format the values if they are numbers
    if isinstance(sales_wsg_percent, (int, float)):
        sales_wsg_percent = format(sales_wsg_percent, '.10f')
    
    if isinstance(net_sales, (int, float)):
        net_sales = format(net_sales, '.10f')
        
    if isinstance(sales_production_wsg, (int, float)):
        sales_production_wsg = format(sales_production_wsg, '.10f')
    
    # Get current timestamp for Created field
    created_timestamp = datetime.now()
    
    print(f"Formatted values - Sales WSG%: {sales_wsg_percent}, Net Sales: {net_sales}, Sales-Production WSG: {sales_production_wsg}")
    print(f"Timestamp: {created_timestamp}")
    
    # Check if the new column exists, and add it if it doesn't
    # try:
    #     cursor.execute("""
    #         ALTER TABLE wsg_goal_tracker 
    #         ADD COLUMN IF NOT EXISTS `Sales-Production WSG% YTD` VARCHAR(50)
    #     """)
    #     connection.commit()
    #     print("✅ Verified/added 'Sales-Production WSG% YTD' column")
    # except Error as e:
    #     print(f"Error adding column: {e}")
    #     connection.rollback()
    
    # Insert into the database with the new column
    try:
        cursor.execute("""
            INSERT INTO wsg_goal_tracker 
            (`Sales-Sales WSG% YTD`, `Net-sales`, `Sales-Production WSG% YTD`, Created)
            VALUES (%s, %s, %s, %s)
        """, (sales_wsg_percent, net_sales, sales_production_wsg, created_timestamp))
        
        connection.commit()
        print("✅ WSG Goal Tracker data inserted successfully.")
        
        # Verify the insertion by retrieving the last record
        cursor.execute("SELECT * FROM wsg_goal_tracker ORDER BY id DESC LIMIT 1")
        last_record = cursor.fetchone()
        print(f"Last record in database: {last_record}")
    except Error as e:
        print(f"Error inserting data: {e}")
        connection.rollback()
    finally:
        cursor.close()
        connection.close()
        print("🔒 MySQL connection closed.")

def extract_all_data():
    task_map = {
        "extract_and_insert": [
            ('Marketing', 'value'),
            ('Sales - Sales', 'sales_value'),
            ('Sales - Production', 'prod_sale'),
            ('Sales - Service ', 'service_sale'),
            ('Production - Production', 'prod_prod'),
            ('Service - Production', 'service_prod'),
        ],
        "extract_sales_teams": [
            ('Sales Team', 'team_name','team','employee'),
            ('VTC History', 'vtc_history','date','vtc'),
        ],
        "extract_and_insert_goals_v3": [
           ('Marketing', 'leads_nd', 'AC', 'AN'),
            ('Sales - Sales', 'issued_unique', 'AC', 'AN'),
            ('Sales - Sales', 'quoted', 'AP', 'BA'),
            ('Sales - Sales', 'sold', 'BC', 'BN'),
            ('Sales - Sales', 'canceled', 'BP', 'CA'),
            ('Sales - Sales', 'net', 'CC', 'CN'),
            ('Sales - Sales', 'net_sale', 'CP', 'DA'),
            ('Sales - Sales', 'adl', 'DC', 'DO'),
            ('Sales - Service ', 'sale_service', 'AC', 'AN'),
            ('Sales - Service ', 'sale_service_net_amt', 'AP', 'BA'),
        ],
    }

    for sheet, col in task_map['extract_and_insert']:
        extract_and_insert(sheet, col)

    for sheet, col, col_1, col_2 in task_map['extract_sales_teams']:
        extract_sales_teams(sheet, col, col_1, col_2)

    for sheet, col, start, end in task_map['extract_and_insert_goals_v3']:
        extract_and_insert_goals_v3(sheet, col, start, end)
    
    # Add the new function call
    extract_wsg_goal_tracker()

if __name__ == "__main__":
    extract_all_data()